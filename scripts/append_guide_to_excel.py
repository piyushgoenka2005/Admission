import sys
from pathlib import Path

from openpyxl import load_workbook


def normalize(name: str) -> str:
    return "".join(ch.lower() for ch in name if ch.isalnum())


def main() -> int:
    if len(sys.argv) != 6:
        print("Usage: append_guide_to_excel.py <xlsx_path> <name> <division> <reporting_officer> <email>", file=sys.stderr)
        return 1

    workbook_path = Path(sys.argv[1])
    name = sys.argv[2].strip()
    division = sys.argv[3].strip()
    reporting_officer = sys.argv[4].strip()
    email = sys.argv[5].strip()

    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    workbook = load_workbook(workbook_path)
    sheet = workbook.active

    email_column = 5
    email_header = sheet.cell(row=1, column=email_column).value
    if not email_header:
        sheet.cell(row=1, column=email_column).value = "Guide_Email"

    target_row = None
    normalized_target = normalize(name)

    for row in range(2, sheet.max_row + 1):
        existing_name = sheet.cell(row=row, column=1).value
        if existing_name and normalize(str(existing_name)) == normalized_target:
            target_row = row
            break

    if target_row is None:
        target_row = sheet.max_row + 1

    sheet.cell(row=target_row, column=1).value = name
    sheet.cell(row=target_row, column=2).value = division
    sheet.cell(row=target_row, column=3).value = reporting_officer
    if sheet.cell(row=target_row, column=4).value is None:
        sheet.cell(row=target_row, column=4).value = ""
    sheet.cell(row=target_row, column=email_column).value = email

    workbook.save(workbook_path)
    print(f"Saved guide '{name}' to {workbook_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
